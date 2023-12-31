import docx
from docx.shared import Cm

import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.drawing.image import Image
#from openpyxl.utils import column_index_from_string

from datetime import date

import jwt
from jwt import encode, decode
from datetime import datetime, timedelta

import psycopg2
conn = psycopg2.connect(dbname='emk', user='emk_u', password='cdjkjxm', host='127.0.0.1')
cursor =  conn.cursor()

def toFloat(val):
    if (type(val) == type(" ")) and (',' in  val):
        val = val.replace(',', '.')
    return float(val)

def ch20(blok):
    mA = ''
    bus= ''
    
    for key in ['Э19', 'Э110',  'Э1S2']: 
        if key in blok:
            bus = 'Profibus DP'
    for key in ['Э23', 'Э24', 'Э25', 'Э26']:
        if key in blok:
            bus = 'RS485 Profibus'
    for key in ['Э14', 'Э15', 'Э18', 'Э01', 'Э1S1']: 
        if key in blok:
            bus = 'RS485 Modbus' 
    for key in ['Э16', 'Э17']:
        if key in blok:
            if bus != '':
                mA = ', 4-20мА'
            else:
                mA = '4-20мА'
    return bus + mA

def sql_format(res):
    ans=[]
    for a in res:
        a=str(a[0])
        if (a not in ans) and (a != 'None'):
            ans.append(a)
    
    try:
        ans.sort(key=float)
    except:
        pass
    return ans



class mk_DOCX():
    def __init__(self, elpod = "", list0 = [], list1 = [], list2 = [], list3 = [], list4 = [], list5 = []):
        self.elpod = elpod
        self.list0 = list0
        self.list1 = list1
        self.list2 = list2
        self.list3 = list3
        self.list4 = list4
        #jsn6[5] -> list5[5] 660 B - доп опция для шахтного исполения
        self.list5 = list5

    def ep4(self):
        mA=ch20(self.list3[0])
        
        names1 = ["Тип арматуры", "Маркировка", "Завод-изготовитель",  "Требуемое время закрытия", "Максимальный крутящий момент", "Присоединение к приводу"]
        ans1 = self.list1[:-4]
        
        if self.list1[6] == "РН":
            names1 += ["Количество оборотов для поворота на 90° "]
            ans1 += [self.list1[9]]
        else:
            names1 += ["Количество оборотов для закрытия арматуры "]
            ans1 += [self.list1[9]+""]
        
        

        names2 = ['Исполнение по назначению', 'Режим работы', 'Степень защиты от проникновения пыли и влаги', 'Вращение выходного вала при закрывании', 'Температура окружающей среды (Климат)'] 
        names2 += ['Тип блока управления ', 'Сигнал дистанционного управления', 'Тип блока концевых выключателей (без встроенного пускателя)', 'Механический указатель сигнализации положения', 'Сигнализация положения', 'Сигнал «Момент» 4-20 мА', 'Дублирование шины RS485'] 
        
        ans2 = self.list2 + [self.list3[0], self.list3[2]+mA] + self.list3[3:-4] 

        if self.list1[6] == "РН":
            names2 +=["Тип РН", "Номинальный крутящий момент на выходном валу редуктора, Н·м ", "Диапазон крутящих моментов на выходном валу редуктора, Н·м ", "Тип фланца по ISO 5211 для присоединения редуктора к арматуре "]
            ans2+=[self.list5[0], self.list5[2], self.list5[3], self.list5[4]]

        names2 += ["Электрическое подключение", 'Защитный колпак', 'Цвет окраски', 'Напряжение питания электродвигателя', 'Количество эл./приводов', 'Дополнительные опции', 'Дополнительные требования']
        ans2 += [ f"« {self.list4[0]}» - {self.elpod}", self.list4[1], self.list4[2], self.list5[5], self.list0[-2]] + self.list4[-2:]

        ans3 = self.list0[:-2]

        return {"names1" : names1, "names2" : names2, "ans1" : ans1, "ans2" : ans2, "ans3" : ans3}
    
    def epn(self):
        mA=ch20(self.list1[1])
        names1 = ["Тип арматуры", "Маркировка", "Завод-изготовитель",  "Требуемое время закрытия", "Максимальный крутящий момент", "Присоединение к приводу"]
        ans1 = self.list1[:-4]
        names1.append("Количество оборотов (угол поворота) для закрытия арматуры")
        ans1.append(self.list1[-1])
        ans1[3] = self.list1[7]

        for i in range(len(names1)):
            names1[i] = str(names1[i])
            ans1[i] = str(ans1[i])
        
        names2 = ["Исполнение по назначению", "Режим работы", "Степень защиты от проникновения пыли и влаги", "Вращение выходного вала при закрывании", "Температура окружающей среды (Климат)"] 
        names2 += ["Тип блока управления", "Сигнал дистанционного управления", "Тип блока концевых выключателей (без встроенного пускателя)", "Сигнализация положения", "Сигнал «Момент» 4-20 мА", "Дублирование шины RS485"] 
        names2 += ["Электрическое подключение", "Цвет окраски", "Напряжение питания электродвигателя", "Количество эл./приводов", "Дополнительные опции", "Дополнительные требования"]
        
        ans2 = self.list2 + [self.list3[0], self.list3[2]+mA, self.list3[3]] + self.list3[5:-4] + [ f"« {self.list4[0]}» - {self.elpod}", self.list4[2], self.list5[5], self.list0[-2]] + self.list4[-2:]

        ans3 = self.list0[:-2]

        return {"names1" : names1, "names2" : names2, "ans1" : ans1, "ans2" : ans2, "ans3" : ans3}

    def vimu(self):
        mA=ch20(self.list3[0])
        names2 = ['Маркировка', 'Исполнение по взрывзащите', 'Контроль положения и крутящего момента (усилия) на выходном звене привода', 'Степень защиты от проникновения пыли и влаги', 'Способ включения двигателя привода ', 'Температура окружающей среды (Климат)'] 
        ans2 = [self.list1[1]] + self.list2
        names2 += ['Тип блока управления ', 'Сигнал дистанционного управления', 'Сигнализация положения', 'Сигнал «Момент» 4-20 мА', 'Дублирование шины RS485', 'Промежуточные выключатели', 'Моментные выключатели', 'Концевые выключатели'] 
        ans2 += [self.list3[0], self.list3[2]+mA] + self.list3[5:-1]
        names2 += ["Электрическое подключение", 'Цвет окраски', 'Специальное исполнение', 'Количество', 'Дополнительные опции', 'Дополнительные требования']
        ans2 += [f'«{self.list5[0]}» - {self.elpod}', self.list4[2], self.list5[1], self.list0[4], self.list4[4], self.list4[5]]

        ans3 = self.list0[:-2]

        return {"names1" : [], "names2" : names2, "ans1" : [], "ans2" : ans2, "ans3" : ans3}
    
    def classic(self):
        names1 = ["Тип арматуры", "Маркировка", "Завод-изготовитель",  "Количество оборотов на закрытие", "Максимальный крутящий момент", "Присоединение к приводу"]
        ans1 = self.list1[:-4]
        ans1[3] = self.list1[-1]

        names2 = ['Исполнение по назначению', 'Режим работы', 'Степень защиты от проникновения пыли и влаги', 'Вращение выходного вала при закрывании', 'Температура окружающей среды (Климат)']
        ans2 = self.list2

        names2 += ['Тип блока концевых выключателей (без встроенного пускателя)', "Электрическое подключение", 'Напряжение питания электродвигателя', 'Количество эл./приводов', 'Дополнительные опции', 'Дополнительные требования']
        ans2 += [self.list3[3], f"{self.elpod}", "380 B   3 фазы", self.list0[-2]] + self.list4[-2:]

        #ans2[-6] = self.list5[1]
        ans3 = self.list0[:-2]

        return {"names1" : names1, "names2" : names2, "ans1" : ans1, "ans2" : ans2, "ans3" : ans3}
        

    
class mk_XL():
    def __init__(self, ID, mark, jsn = [[], [], [], [], [], [], [], []]):
        self.jsn = jsn
        self.ID = ID
        self.mark = mark
    
    def ep(self):
        a = self.jsn
        ID =  self.ID

        mark = self.mark
        print(mark)
        if mark[:3] == "ЭП4": 
            wb = load_workbook('ep4TKP.xlsx')
        elif mark[:3] == "ЭПН":
            wb = load_workbook('epnTKP.xlsx')
        sheet = wb['Формат ТКП']

        if 'М1' == a[3][0][:2]:
            M = DB(a[3][0])
            M1 = M.get_M1(a[1][8], a[1][4], a[1][7], a[1][9])
            a[3][1] += f"Верхний предел настройки путевых выключателей в оборотах выходного вала- {M1[1:]}."

        #Лицевая часть
        current_date = date.today()
        keys = ["I7", "C9", "C10", "C12", "I12", "C48"]
        ans = [ID, f"Организация: {a[0][0]}   ФИО представителя: {a[0][1]}   Тел.: {a[0][2]}   email: {a[0][3]}", "", mark, str(a[0][-2]), str(f"{current_date.day}.{current_date.month}.{current_date.year}")]

        #Техническая часть
        for i in range(16, 40):
            k = 'G' + str(i)
            keys.append(k)
        red_info = ""
        if a[6][0] != "?":
            red_info = f"{a[6][0]} (на выходном валу: номинальный крутящий момент редуктора {a[6][2]} Н·м, диапазон крутящих моментов:  {a[6][3]} Н·м. Тип фланца по ISO 5211 для присоединения редуктора к арматуре: {a[6][4]})."
        ans = ans + [
            f'{a[1][0]}', #4 "Поворотный/многооборотный",
            
            f'{a[2][1]}', #5 "Запорный/ регулирующий",
            f'{a[2][0]}', #6 "взрывозащита/Рудничное/АЭС и тд",
            f'{a[1][5]}', #7 "Тип присоединительного фланца ",
            f'{a[1][4]}', #8 "Верхний предел настройки крутящего момента",
            f'{a[1][7]}', #9 "Частота вращения выходного вала ",

            f'{a[3][0]}', #10.1 "Наименование блока управления"
            f'{a[3][1]}', #10.2 "Тип управления: Расшифровка дается по всем возможным параментрам, протоколы, возможне сигналы",

            f'{a[2][4]}', #11 "Температурное исполнение: температурный режим, климат,  в укрытии или под навесом",
            f'{a[6][1]}', #12 "Тип присоединения выходного вала привода с валом арматуры",

            f'{a[2][3]}', #13 "Вращение выходного вала при закрывании",
            f'{a[2][2]}', #14 "Уровень пылевлагозащиты",
            f'{a[4][2]}', #15 "Цвет окраски",

            f'«{a[4][0]}» - {a[5][1]}', #16 "Электрическое подключение",
            f'{a[3][10]}', #17 "Концевые выключатели (реле)",
            f'{a[3][8]}', #18 "Промежуточные выключатели (опция)",
            f'{a[3][9]}', #19 "Моментные выключатели"
            f'{a[3][4]}', #20 "Механический указатель положения",
            f'{a[3][2]} \n {a[3][5]}', #21 "Дистанционный указатель положения. Датчики обратной связи)",
            f'{a[4][3]}', #22 "Ручной селектор/переключение местного дистанционного управления",
            f'{a[3][11]}', #23 "Монтаж блока управления ",

            f'{a[5][2]}', #24 "Требования по функциональной безопасности SIL",
            f'{a[5][3]}', #25 "Специальное исполнение",
            f'{red_info} \n {a[4][5]}', #26 "Особые требования ",
        ]

        #Характеристики двигателя
        for i in range(42, 48):
            k = 'G' + str(i)
            keys.append(k)

        #Маркировка
        mark = a[1][1][:3]
        #конструктивная схема
        sh = a[1][8]
        #крутящий момент
        Mom = a[1][4]
        #частота вращения
        V = a[1][7]
        #фланец
        flc = a[1][5]
        if mark == "ЭП4":
            bd = DataBase()
            aaa = bd.get_engin(mark, sh, Mom, V, flc)
            print(aaa)
            aaa[4] = a[6][5]
        if mark == "ЭПН":
            bd = DataBase()
            U = a[6][5]
            if U[:2] == '24':
                U='24'
                fz=''
            elif a[6][5][:3] == '380':
                U = '380'
                fz = '3'
            elif a[6][5][:3] == '220':
                U='220'
                fz = a[6][5][6] 
            else:
                U = ''
                fz =''
            aaa = bd.get_engin(mark, sh, Mom, V, flc, U, fz)

        
        ans = ans + aaa

        for i in range(len(keys)):
            if ans[i] != None:
                sheet[keys[i]].value = ans[i]

        sheet["G40"].value = f"Не более {ans[-1]} кг"

        return wb
    
    def vimu(self):
        a = self.jsn

        wb = load_workbook('vimuTKP.xlsx')
        sheet = wb['Формат ТКП']

        #Лицевая часть
        current_date = date.today()
        keys = ["I7", "C9", "C10", "C12", "I12", "C31"]
        mark = a[1][1]
        ans = [self.ID, f"Организация: {a[0][0]}   ФИО представителя: {a[0][1]}   Тел.: {a[0][2]}   email: {a[0][3]}", "", self.mark, str(a[0][-2]), str(f"{current_date.day}.{current_date.month}.{current_date.year}")]

        #Техническая часть
        for i in range(16, 31):
            k = 'G' + str(i)
            keys.append(k)

        ans = ans + [
            f'{a[1][0]}', #4 "Исполнение",
            f'{a[2][0]}', #5.1 "Наименование блока управления"
            f'{a[2][1]}', #5.2 "Тип управления: Расшифровка дается по всем возможным параментрам, протоколы, возможне сигналы",
            f'{a[1][4]}', #6 "Температурное исполнение: температурный режим, климат,  в укрытии или под навесом",

            f'{a[1][1]}', #7 "Контроль положения и крутящего момента (усилия) на выходном звене привода"
            f'{a[1][3]}', #8 "Способ включения двигателя привода"

            f'{a[1][2]}', #9 "Уровень пылевлагозащиты",
            f'{a[3][1]}', #10 "Цвет окраски",

            f'«{a[3][0]}» - {a[3][4]}', #11 "Электрическое подключение",
            f'{a[2][10]}', #12 "Концевые выключатели (реле)",
            f'{a[2][9]}', #13 "Моментные выключатели"
            f'{a[2][8]}', #14 "Промежуточные выключатели (опция)",
            
            f'{a[2][2]} \n {a[2][5]}', #15 "Дистанционный указатель положения / датчики обратной связи)",

            f'{a[3][5]}', #16 "Специальное исполнение",

            f'{a[3][3]}', #17 "Особые требовнаия",
        ]

        for i in range(len(keys)):
            if ans[i] != None:
                sheet[keys[i]].value = ans[i]

        return wb
    
    def classic(self):
        a = self.jsn
        ID =  self.ID

        mark = self.mark
        print(mark)
        wb = load_workbook('classicTKP.xlsx')
        sheet = wb['Формат ТКП']

        #Лицевая часть
        current_date = date.today()
        keys = ["I7", "C9", "C10", "C12", "I12", "C42"]
        ans = [ID, f"Организация: {a[0][0]}   ФИО представителя: {a[0][1]}   Тел.: {a[0][2]}   email: {a[0][3]}", "", mark, str(a[0][-2]), str(f"{current_date.day}.{current_date.month}.{current_date.year}")]

        #Техническая часть
        for i in range(16, 34):
            k = 'G' + str(i)
            keys.append(k)
        bd = DB("Классика")
        
        ans = ans + [
            f'{a[1][0]}', #4 "Поворотный/многооборотный",
            
            f'{a[2][1]}', #5 "Запорный/ регулирующий",
            f'{a[2][0]}', #6 "взрывозащита/Рудничное/АЭС и тд",
            f'{a[1][8]}', #7 "Модернизация",
            f'{a[1][5]}', #8 "Тип присоединительного фланца ",
            f'{a[1][4]}', #9 "Верхний предел настройки крутящего момента",
            f'{a[1][7]}', #10 "Частота вращения выходного вала"
            f'{a[1][9]}', #11 "Количество оборотов на закрытие",

            f'{a[2][4]}', #12 "Температурное исполнение: температурный режим, климат,  в укрытии или под навесом",
            f'{a[6][1]}', #13 "Тип присоединения выходного вала привода с валом арматуры",

            f'{a[2][3]}', #14 "Вращение выходного вала при закрывании",
            f'{a[2][2]}', #15 "Уровень пылевлагозащиты",

            f'{a[5][1]}', #16 "Электрическое подключение",
            f'{a[3][3]}',#17 "Тип блока концевых выключателей",
            f'{a[3][10]}', #18 "Концевые выключатели (реле)",
            f'{a[3][9]}', #19 "Моментные выключатели"
            f'{a[3][8]}',#20 "Промежуточные выключатели (опция)",
            f'{a[4][5]}',#21 "Особые требования",
        ]

        #Характеристики двигателя
        for i in range(36, 42):
            k = 'G' + str(i)
            keys.append(k)

        #Маркировка
        bd = DataBase()
        if a[1][8] == 'Отсутствует':
            mod = 0
        else: 
            mod = a[1][8]
        aaa = bd.get_class_engin(isp = a[6][0], flc = a[1][5], mod = mod, nom = a[5][4])
        ans = ans + aaa

        for i in range(len(keys)):
            if ans[i] != None:
                sheet[keys[i]].value = ans[i]

        sheet["G34"].value = f"Не более {ans[-1]} кг"

        return wb



class DB():
    def __init__(self, mark):
        self.mark = mark
    
    def get_engin(self, mark, sh, Mom, V, flc, U=""):
        WB = load_workbook('BD.xlsx')
        Sheet = WB['Связь']

        print("Подбор двигателя для ЭП")
            
        ansi = []
        for i in range(2, 1500):
            if ((mark[:3] == str(Sheet[f"A{i}"].value)) and str(int(Sheet[f"B{i}"].value)) == str(int(sh))) and (str(float(Sheet[f"D{i}"].value)) == str(float(Mom))) and (str(float(Sheet[f"E{i}"].value)) == str(float(V))) and (str(Sheet[f"C{i}"].value) == str(flc)):
                if (mark[:3] == 'ЭПН'):
                #print(i)
                    u = Sheet[f"L{i}"].value
                    UU = f"{u}"
                    if (UU[:2] == str(U)[:2]):
                        ansi.append(i)
                else:
                    ansi.append(i)


                    
            if ansi != []:
                i = ansi[0]
                ans = [str(Sheet[f"H{i}"].value) + " кВт", str(Sheet[f"I{i}"].value) + " кВт", str(Sheet[f"J{i}"].value) + " А", str(Sheet[f"K{i}"].value) + " А", str(Sheet[f"L{i}"].value) + " В", str(Sheet[f"M{i}"].value), str(Sheet[f"F{i}"].value)]
            else:
                ans = "Не удалось подобрать"
        return ans

    def get_params(self, mark, flc_type = "", Hm = "", N = "", V = "", sh = "", flc = "", fz = ""):
        ans = []
        WB = load_workbook('BD.xlsx')
        Sheet = WB['Электропривода']
        sheet = WB['Связь']



        print("Подбор параметров для ЭП")

        if flc_type == "" or flc_type == None:
            ans = "Укажите тип фланца"

        elif Hm == "" or Hm ==None:
            ans = []
            #print("Крутящий момент не задан")
            for i in range(2, 1020):
                if (str(Sheet[f"A{i}"].value) == mark) and (str(Sheet[f"G{i}"].value) == str(flc_type)):
                    if Sheet[f"D{i}"].value not in ans:
                        ans.append(Sheet[f"D{i}"].value)
        
        elif N == "" or N == None:
            ans = []
            #print("Крутящий момент не задан")
            for i in range(2, 1500):
                N = sheet[f'H{i}'].value
                if (str(Sheet[f"A{i}"].value) == mark) and (str(Sheet[f"G{i}"].value) == str(flc_type)) and (str(Sheet[f"D{i}"].value) == str(Hm)):
                    if N not in ans:
                        ans.append(N)

        elif V == "" or V ==None:
            ans = []
            #print("Частота не задана")
            for i in range(2, 1020):
                if (Sheet[f"A{i}"].value == mark) and (str(Sheet[f"G{i}"].value) == str(flc_type)) and (str(Sheet[f"D{i}"].value) == str(Hm)) and ((str(N) == str(sheet[f'H{i}'].value)) or (str(N) == "0")):
                    if  Sheet[f"E{i}"].value not in ans:
                        ans.append(Sheet[f"E{i}"].value)

        elif sh == "" or sh == None:
            ans = []
            #print("Конструктивная схема не задана")
            for i in range(2, 1020):
                if (Sheet[f"A{i}"].value == mark) and (str(Sheet[f"G{i}"].value) == str(flc_type)) and (str(Sheet[f"D{i}"].value) == str(Hm)) and (str(Sheet[f"E{i}"].value) == str(V)) and ((str(N) == str(sheet[f'H{i}'].value)) or (str(N) == "0")):
                    if Sheet[f"B{i}"].value not in ans:
                        ans.append(Sheet[f"B{i}"].value)
            
        elif flc == "" or flc == None:
            ans = []
            #print("Все парметры указаны верно")
            for i in range(2, 1020):
                if (Sheet[f"A{i}"].value == mark) and (str(Sheet[f"G{i}"].value) == str(flc_type)) and (str(Sheet[f"D{i}"].value) == str(Hm)) and (str(Sheet[f"E{i}"].value) == str(V)) and (str(Sheet[f"B{i}"].value) == str(sh)) and ((str(N) == str(sheet[f'H{i}'].value)) or (str(N) == "0")):
                    if Sheet[f"C{i}"].value not in ans:
                        ans.append(Sheet[f"C{i}"].value)

        elif (flc_type != "" and Hm != "" and V != "" and sh != "" and flc != ""):
            sheet = WB["Связь"]
            for i in range(2, 1500):
                if (sheet[f"A{i}"].value == mark) and (str(sh) == str(sheet[f'B{i}'].value)) and (str(flc) == str(sheet[f'c{i}'].value)) and (str(Hm) == str(sheet[f'D{i}'].value)) and (str(V) == str(sheet[f'E{i}'].value)) and ((str(N) == str(sheet[f'H{i}'].value)) or (str(N) == "0")):
                    if str(sheet[f"L{i}"].value) + "B " + str(sheet[f"M{i}"].value) + " фаз(ы) " not in ans:
                        if str(sheet[f"L{i}"].value) == "220" or "380":
                            ans.append(str(sheet[f"L{i}"].value) + "B " + str(sheet[f"M{i}"].value) + " фаз(ы) ")
                        elif str(sheet[f"L{i}"].value) == "24":
                            ans.append(str(sheet[f"L{i}"].value) + "B ")

        else:
            ans = "Ошибка выбора параметров"
        
        return ans
    
    def get_RN(self, rn = "", flc = "", Hm = "", V = "", t="", sh = ""):
        ans = []
        WB = load_workbook('BD.xlsx')
        Sheet = WB['Электропривода']
        
        

        print("Подбор параметров для ЭП4 с РН")

        if rn == "":
            for i in range(2, 568):
                rn = Sheet[f"I{i}"].value
                if (rn != "") and (rn not in ans):
                    ans.append(rn)
        elif flc == "":
            for i in range(2, 568):
                RN = Sheet[f"I{i}"].value
                flc = Sheet[f"C{i}"].value
                if (flc != "") and (flc not in ans)  and (rn == RN):
                    ans.append(flc)
        elif Hm == "":
            for i in range(2, 568):
                RN = Sheet[f"I{i}"].value
                Flc = Sheet[f"C{i}"].value
                Hm = Sheet[f"D{i}"].value
                if (Hm != "") and (Hm not in ans) and (flc == Flc) and (rn == RN):
                    ans.append(Hm)
        elif V == "":
            for i in range(2, 568):
                RN = Sheet[f"I{i}"].value
                Flc = Sheet[f"C{i}"].value
                HM = Sheet[f"D{i}"].value
                V = Sheet[f"E{i}"].value
                if (V != "") and (V not in ans) and (str(Hm) == str(HM)) and (flc == Flc) and (rn == RN):
                    ans.append(V)
        elif t == "":
            for i in range(2, 568):
                Flc = Sheet[f"C{i}"].value
                HM = Sheet[f"D{i}"].value
                v = Sheet[f"E{i}"].value
                RN = Sheet[f"I{i}"].value
                t = Sheet[f"J{i}"].value
                if (t != "") and (t not in ans) and (str(V) == str(v)) and (str(Hm) == str(HM)) and (flc == Flc) and (rn == RN) :
                    ans.append(t)
        elif sh == "":
            for i in range(2, 568):
                RN = Sheet[f"I{i}"].value
                Flc = Sheet[f"C{i}"].value
                HM = Sheet[f"D{i}"].value
                v = Sheet[f"E{i}"].value
                RN = Sheet[f"I{i}"].value
                T = Sheet[f"J{i}"].value
                sh = Sheet[f"B{i}"].value
                if (sh != "") and (sh not in ans) and (str(V) == str(v)) and (str(Hm) == str(HM)) and (flc == Flc) and (int(t) == T) and (rn == RN):
                    ans.append(sh)
        return ans
    
    def get_classic(self, isp = "", flc = "", ob = "", V = "", Hm = "", bkw = "", elpod = "", nom = ""):
        ans = []
        isp = str(isp)
        flc = str(flc)
        nom = str(nom) 
        ob = str(ob)
        V = str(V)
        Hm = str(Hm)
        BKW = str(bkw)
        ElPod = str(elpod)
        WB = load_workbook('BD.xlsx')
        Sheet = WB['Классика']

        print("Подбор параметров для классики")

        if isp == "":
            for i in range(2, 417):
                isp = Sheet[f"B{i}"].value
                if (isp != "") and (isp not in isp):
                    ans.append(isp)
        elif flc == "":
            for i in range(2, 417):
                ISP = Sheet[f"B{i}"].value
                flc = Sheet[f"C{i}"].value
                if (flc != "") and (flc not in ans)  and (isp == ISP):
                    ans.append(flc)
        elif ob == "":
            for i in range(2, 417):
                ISP = Sheet[f"B{i}"].value
                FLC = Sheet[f"C{i}"].value
                ob = Sheet[f"H{i}"].value
                if (ob != "") and (ob not in ans)  and (isp == ISP) and (flc == FLC):
                    ans.append(ob)
        elif V == "":
            for i in range(2, 417):
                ISP = Sheet[f"B{i}"].value
                FLC = Sheet[f"C{i}"].value
                OB = Sheet[f"H{i}"].value
                V = Sheet[f"G{i}"].value
                if (V != "") and (V not in ans)  and (isp == ISP) and (flc == FLC) and (int(ob) == int(OB)):
                    ans.append(V)
        elif Hm == "":
            for i in range(2, 417):
                ISP = Sheet[f"B{i}"].value
                FLC = Sheet[f"C{i}"].value
                OB = Sheet[f"H{i}"].value
                v = Sheet[f"G{i}"].value
                Hm = Sheet[f"F{i}"].value
                if (Hm != "") and (Hm not in ans)  and (isp == ISP) and (flc == FLC) and (int(ob) == int(OB)) and (float(v) == float(V)):
                    ans.append(Hm)
        elif bkw == "":
            for i in range(2, 417):
                ISP = Sheet[f"B{i}"].value
                FLC = Sheet[f"C{i}"].value
                OB = Sheet[f"H{i}"].value
                v = Sheet[f"G{i}"].value
                HM = Sheet[f"F{i}"].value
                bkw = Sheet[f"P{i}"].value
                if (bkw != "") and (bkw not in ans)  and (isp == ISP) and (flc == FLC) and (int(ob) == int(OB)) and (float(v) == float(V)) and (float(Hm) == float(HM)):
                    ans.append(bkw)
        elif elpod == "":
            for i in range(2, 417):
                ISP = Sheet[f"B{i}"].value
                FLC = Sheet[f"C{i}"].value
                OB = Sheet[f"H{i}"].value
                v = Sheet[f"G{i}"].value
                HM = Sheet[f"F{i}"].value
                BKW = Sheet[f"P{i}"].value
                elpod = Sheet[f"Q{i}"].value
                if (elpod != "") and (elpod not in ans)  and (isp == ISP) and (flc == FLC) and (int(ob) == int(OB)) and (float(v) == float(V)) and (float(Hm) == float(HM) and (bkw == BKW)):
                    ans.append(elpod)
        else:
            res = []
            for i in range(2, 417):
                ISP = Sheet[f"B{i}"].value
                FLC = Sheet[f"C{i}"].value
                OB = Sheet[f"H{i}"].value
                v = Sheet[f"G{i}"].value
                HM = Sheet[f"F{i}"].value
                BKW = Sheet[f"P{i}"].value
                ElPod = Sheet[f"Q{i}"].value
                nom = Sheet[f"E{i}"].value
                if (nom != "") and (i not in res)  and (isp == ISP) and (flc == FLC) and (int(ob) == int(OB)) and (float(v) == float(V)) and (float(Hm) == float(HM) and (bkw == BKW) and (elpod == ElPod)):
                    res.append(i)
            for i in res:
                md = Sheet[f"D{i}"].value
                nm = Sheet[f"E{i}"].value
                ans.append([md, nm])
        
        

        return ans
    
    def get_class_engin(self, isp = "", flc = "", mod = "", nom = ""):
        WB = load_workbook('BD.xlsx')
        Sheet = WB['Классика']
        print("Подбор двигателя для классики")


        ansi = []
        for i in range(2, 416):
            ISP = Sheet[f"B{i}"].value
            FLC = Sheet[f"C{i}"].value
            MOD = Sheet[f"D{i}"].value
            NOM = Sheet[f"E{i}"].value
            if MOD == "Отсутствует":
                MOD = 0
            if ((str(isp) == str(ISP)) and (str(flc) == str(FLC)) and (int(mod) == int(MOD)) and (int(nom) == int(NOM)) ):
                ansi.append(i)
                    
            if ansi != []:
                for i in ansi:
                    ans = [str(Sheet[f"J{i}"].value) + " кВт", str(Sheet[f"K{i}"].value) + " кВт", str(Sheet[f"L{i}"].value) + " А", str(Sheet[f"M{i}"].value) + " А", str(Sheet[f"N{i}"].value) + " В", str(Sheet[f"O{i}"].value), str(Sheet[f"I{i}"].value)]
                    print(ans)
            else:
                ans = "Не удалось подобрать"
        return ans
    
    def get_mark(self, arr=["", "", "", "", "", ""]):
        ans = []
        Mark = arr[0]
        if Mark == 'ЭП4':
            WB = load_workbook('BD.xlsx')
            Sheet = WB['Электропривода']
            flc = str(arr[1])
            Hm = str(arr[2])
            V = str(arr[3])
            type_flc = str(arr[4])
            sh = str(arr[5])
            print(f"Подбор маркировки для {mark}:")
            print(f"[фланец, момент, частота, тип фланца, схема]")
            print(arr)
            if flc == "":
                for i in range(2, 568):
                    FLC = str(Sheet[f"C{i}"].value)
                    if (FLC not in ans):
                        ans.append(FLC)
            elif Hm == "":
                for i in range(2, 568):
                    FLC = str(Sheet[f"C{i}"].value)
                    HM = str(Sheet[f"D{i}"].value)
                    if (HM not in ans) and (HM != "") and (flc == FLC):
                        ans.append(HM)
            elif V == "":
                for i in range(2, 568):
                    FLC = str(Sheet[f"C{i}"].value)
                    HM = str(Sheet[f"D{i}"].value)
                    v = str(Sheet[f"E{i}"].value)
                    if (v not in ans) and (v != "") and (Hm == HM) and (flc == FLC):
                        ans.append(v)
            elif type_flc == "":
                for i in range(2, 568):
                    FLC = str(Sheet[f"C{i}"].value)
                    HM = str(Sheet[f"D{i}"].value)
                    v = str(Sheet[f"E{i}"].value)
                    TFlc = str(Sheet[f"G{i}"].value)
                    if (TFlc not in ans) and (TFlc != "") and (V == v) and (Hm == HM) and (flc == FLC):
                        ans.append(TFlc)
            elif sh == "":
                for i in range(2, 568):
                    FLC = str(Sheet[f"C{i}"].value)
                    HM = str(Sheet[f"D{i}"].value)
                    v = str(Sheet[f"E{i}"].value)
                    TFlc = str(Sheet[f"G{i}"].value)
                    SH = str(Sheet[f"B{i}"].value)
                    if (SH not in ans) and (SH != "") and (type_flc == TFlc) and (V == v) and (Hm == HM) and (flc == FLC):
                        ans.append(SH)


                
        elif Mark == 'ЭПН':
            WB = load_workbook('BD.xlsx')
            Sheet = WB['Электропривода']
            sheet = WB['Связь']
            flc = str(arr[1])
            Hm = str(arr[2])
            V = str(arr[3])
            fz = str(arr[4])
            sh = str(arr[5])
            print(f"Подбор маркировки для {Mark}:")
            print(f"[фланец, моент, частота, тип фланца, схема]")
            print(arr)
            
            if flc == "":
                for i in range(568, 1200):
                    FLC = str(Sheet[f"C{i}"].value)
                    if (FLC not in ans):
                        ans.append(FLC)
            elif Hm == "":
                for i in range(568, 1200):
                    FLC = str(Sheet[f"C{i}"].value)
                    HM = str(Sheet[f"D{i}"].value)
                    if (HM not in ans) and (HM != "") and (flc == FLC):
                        ans.append(HM)
            elif V == "":
                for i in range(568, 1200):
                    FLC = str(Sheet[f"C{i}"].value)
                    HM = str(Sheet[f"D{i}"].value)
                    v = str(Sheet[f"E{i}"].value)
                    if (v not in ans) and (v != "") and (Hm == HM) and (flc == FLC):
                        ans.append(v)
            elif fz == "":
                for i in range(568, 1200):
                    FLC = str(Sheet[f"C{i}"].value)
                    HM = str(Sheet[f"D{i}"].value)
                    v = str(Sheet[f"E{i}"].value)
                    FZ = str(sheet[f"M{i}"].value)
                    if (FZ not in ans) and (FZ != "") and (V == v) and (Hm == HM) and (flc == FLC):
                        ans.append(FZ)
            elif sh == "":
                for i in range(568, 1200):
                    FLC = str(Sheet[f"C{i}"].value)
                    HM = str(Sheet[f"D{i}"].value)
                    v = str(Sheet[f"E{i}"].value)
                    FZ = str(sheet[f"M{i}"].value)
                    SH = str(Sheet[f"B{i}"].value)
                    if (SH not in ans) and (SH != "") and (fz == FZ) and (V == v) and (Hm == HM) and (flc == FLC):
                        ans.append(SH)
        elif Mark == 'ВИМУ':
            WB = load_workbook('BD.xlsx')
            #Sheet = WB['ВИМУ']
        
        return ans

    def get_M1(self, sh, mom, v, ob):
        sh = int(sh)
        mom = int(mom)
        v = float(v)
        ob = float(ob)
        print(f"Подбор блока М1 при схеме - {sh}, моменте - {mom}, частоте - {v}, и оборотах - {ob}.")
        #Sim = ['M', 'L', 'K', 'J', 'I', 'H', 'G', 'F', 'E', 'D'] #слева направо
        Sim = ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M'] #справа налево
        k = 2
        m = 70


        WB = load_workbook("M1.xlsx")
        sheet = WB["M1"]

        for i in range(3, 70):
            SH = sheet[f'A{i}'].value
            Mom = sheet[f'B{i}'].value
            V = float(sheet[f'C{i}'].value) 
            if (sh == SH) and (mom == Mom) and (v == V):
                k = i
                m = i+68

        for s in Sim:
            pos_min = f"{s}{k}"
            pos_max = f"{s}{m}"

            min_ob = sheet[pos_min].value
            max_ob = sheet[pos_max].value
            
            if min_ob == "-":
                 pos_min = f"{s}2"
                 min_ob = sheet[pos_min].value
            if max_ob == "-":
                 pos_max = f"{s}70"
                 max_ob = sheet[pos_max].value
            
            min_ob = toFloat(min_ob)
            max_ob = toFloat(max_ob)

            if min_ob <= ob <= max_ob:
                print(sheet[f"{s}1"].value)
                if sheet[f"{s}1"].value == 2.5: 
                    return ".2,5"
                return "." + str(sheet[f"{s}1"].value)
        
        return ""
    


class DataBase():
    def get_engin(self, mark, sh, Mom, V, flc, U ="", fz =""):
        conn = psycopg2.connect(dbname='emk', user='emk_u', password='cdjkjxm', host='127.0.0.1')
        cursor =  conn.cursor()
        ans = []
        mark = str(mark)
        sh=str(sh)
        Mom=int(Mom)
        V=float(V)
        flc=str(flc)
        request=""
        if U != "":
           request = f'AND U = \'{U}\''
        if str(U) == '220':
            request = request + f"AND fz = \'{fz}\'"
        print(f'SELECT (N, N, Ipusk, Inom, U, fz, Mas) FROM ElectroDrive WHERE Mark=\'{mark}\' AND sh=\'{sh}\' AND Mom=\'{Mom}\' AND V=\'{V}\' AND flc=\'{flc}\' {request};')
        cursor.execute(f'SELECT (N, N, Ipusk, Inom, U, fz, Mas) FROM ElectroDrive WHERE Mark=\'{mark}\' AND sh=\'{sh}\' AND Mom=\'{Mom}\' AND V=\'{V}\' AND flc=\'{flc}\' {request};')
        res = cursor.fetchone()[0].split(',')
        print(res)
        ans = [str(res[0][1:]) + " кВт", str(res[1]) + " кВт", str(res[2]) + " А", str(res[3]) + " А", str(res[4]) + " В", str(res[5]), str(res[6][:-1])]
        cursor.close() # закрываем курсор
        conn.close()
        return ans
    
    def get_params(self, mark, Isp = "", flc_type = "", Mom = "", N = "", V = "", sh = "", flc = "", fz = ""):
        conn = psycopg2.connect(dbname='emk', user='emk_u', password='cdjkjxm', host='127.0.0.1')
        cursor =  conn.cursor()
        ans = []
        mark = str(mark)
        Isp = str(Isp)

        if mark == 'ЭПН':
            flc_type = " IS NULL "
        equal=""
        if mark == 'ЭП4':
            equal = "="


        print(f"Подбор параметров для {mark}")
        if Isp == "":
            cursor.execute(f'SELECT Isp FROM ElectroDrive WHERE Mark=\'{mark}\';')
            res = cursor.fetchall()
            ans = sql_format(res)
            cursor.close() # закрываем курсор
            conn.close()
            return ans
        
        elif flc_type == "":
            
            cursor.execute(f'SELECT flc_type FROM ElectroDrive WHERE Mark=\'{mark}\' AND Isp=\'{Isp}\';')
            res = cursor.fetchall()
            ans = sql_format(res)
            cursor.close() # закрываем курсор
            conn.close()
            return ans
        
        elif Mom == "":
            cursor.execute(f'SELECT Mom FROM ElectroDrive WHERE Mark=\'{mark}\' AND Isp=\'{Isp}\' AND flc_type{equal}{flc_type};')
            res = cursor.fetchall()
            ans = sql_format(res)
            cursor.close() # закрываем курсор
            conn.close()
            return ans
        
        elif N == "" and V == "":
            Mom=int(Mom)
            cursor.execute(f'SELECT N FROM ElectroDrive WHERE Mark=\'{mark}\' AND Isp=\'{Isp}\' AND flc_type{equal}{flc_type} AND Mom={Mom};')
            res = cursor.fetchall()
            ans = sql_format(res)
            cursor.close() # закрываем курсор
            conn.close()
            return ans
        
        elif V == "" and N == "0":
            Mom=int(Mom)
            cursor.execute(f'SELECT V FROM ElectroDrive WHERE Mark=\'{mark}\' AND Isp=\'{Isp}\' AND flc_type{equal}{flc_type} AND Mom={Mom};')
            res = cursor.fetchall()
            ans = sql_format(res)
            cursor.close() # закрываем курсор
            conn.close()
            return ans
        
        elif V == "":
            Mom=int(Mom)
            N=float(N)
            cursor.execute(f'SELECT V FROM ElectroDrive WHERE Mark=\'{mark}\' AND Isp=\'{Isp}\' AND flc_type{equal}{flc_type} AND Mom={Mom} AND N=\'{N}\';')
            res = cursor.fetchall()
            ans = sql_format(res)
            cursor.close() # закрываем курсор
            conn.close()
            return ans
        
        elif sh == "":
            Mom=int(Mom)
            V=float(V)
            response=""
            if N != '0':
                N = float(N)
                response = f" AND N=\'{N}\'"
            print(f'SELECT sh FROM ElectroDrive WHERE Mark=\'{mark}\' AND Isp=\'{Isp}\' AND flc_type{equal}{flc_type} AND Mom=\'{Mom}\' AND V=\'{V}\' {response};')
            cursor.execute(f'SELECT sh FROM ElectroDrive WHERE Mark=\'{mark}\' AND Isp=\'{Isp}\' AND flc_type{equal}{flc_type} AND Mom=\'{Mom}\' AND V=\'{V}\' {response};')
            res = cursor.fetchall()
            ans = sql_format(res)
            cursor.close() # закрываем курсор
            conn.close()
            return ans
        
        elif flc == "":
            Mom=int(Mom)
            V=float(V)
            response=""
            if N != '0':
                N = float(N)
                response = f" AND N=\'{N}\'"
            cursor.execute(f'SELECT flc FROM ElectroDrive WHERE Mark=\'{mark}\' AND Isp=\'{Isp}\' AND flc_type{equal}{flc_type} AND Mom={Mom} AND V=\'{V}\' {response}  AND sh=\'{sh}\';')
            res = cursor.fetchall()
            ans = sql_format(res)
            cursor.close() # закрываем курсор
            conn.close()
            return ans
        
        elif fz == "":
            Mom=int(Mom)
            V=float(V)
            sh=int(sh)
            flc=str(flc)
            response=""
            if N != '0':
                N = float(N)
                response = f" AND N=\'{N}\'"
            cursor.execute(f'SELECT fz FROM ElectroDrive WHERE Mark=\'{mark}\' AND Isp=\'{Isp}\' AND flc_type{equal}{flc_type} AND Mom={Mom} AND V=\'{V}\' {response}  AND sh=\'{sh}\' AND flc=\'{flc}\';')
            res_fz = cursor.fetchall()
            cursor.execute(f'SELECT U, fz FROM ElectroDrive WHERE Mark=\'{mark}\' AND Isp=\'{Isp}\' AND flc_type{equal}{flc_type} AND Mom={Mom} AND V=\'{V}\' {response}  AND sh=\'{sh}\' AND flc=\'{flc}\';')
            
            res = cursor.fetchall()    
            print(res)
            for a in res:
                U=int(a[0])
                if U == 220:
                    a = f"{U} В {a[1]} фаз(а/ы)"
                elif  U == 380:
                    a = "380 В 3 фазы"
                elif U == 24:
                    a = "24 В"
                if a not in ans:
                    ans.append(a)

            cursor.close() # закрываем курсор
            conn.close()
            return ans
    
    def get_RN(self, rn = "", isp = "", flc = "", Mom = "", V = "", t="", sh = ""):
        conn = psycopg2.connect(dbname='emk', user='emk_u', password='cdjkjxm', host='127.0.0.1')
        cursor =  conn.cursor()
        ans=[]
        rn=str(rn)

        print("Подбор ЭП4РН")
        if rn == "":
            cursor.execute(f'SELECT rn FROM ElectroDrive WHERE RN IS NOT NULL;')
            res = cursor.fetchall()
            ans = sql_format(res)
            cursor.close() # закрываем курсор
            conn.close()
            return ans
        
        elif isp == "":
            cursor.execute(f'SELECT isp FROM ElectroDrive WHERE rn=\'{rn}\';')
            res = cursor.fetchall()
            ans = sql_format(res)
            cursor.close() # закрываем курсор
            conn.close()
            return ans
        
        elif flc == "":
            cursor.execute(f'SELECT flc FROM ElectroDrive WHERE rn=\'{rn}\' AND isp = \'{isp}\';')
            res = cursor.fetchall()
            ans = sql_format(res)
            cursor.close() # закрываем курсор
            conn.close()
            return ans
        
        elif Mom == "":
            cursor.execute(f'SELECT Mom FROM ElectroDrive WHERE rn=\'{rn}\' AND flc =\'{flc}\' AND isp = \'{isp}\';')
            res = cursor.fetchall()
            ans = sql_format(res)
            cursor.close() # закрываем курсор
            conn.close()
            return ans
        
        elif V == "":
            print(f'SELECT V FROM ElectroDrive WHERE rn=\'{rn}\' AND flc =\'{flc}\' and Mom=Mom AND isp = \'{isp}\';')
            cursor.execute(f'SELECT V FROM ElectroDrive WHERE rn=\'{rn}\' AND flc =\'{flc}\' and Mom={Mom} AND isp = \'{isp}\';')
            res = cursor.fetchall()
            ans = sql_format(res)
            cursor.close() # закрываем курсор
            conn.close()
            print(ans)
            return ans
        
        elif t == "":
            cursor.execute(f'SELECT t_pov FROM ElectroDrive WHERE rn=\'{rn}\' AND flc =\'{flc}\' and Mom={Mom} AND V=\'{V}\' AND isp = \'{isp}\';')
            res = cursor.fetchall()
            ans = sql_format(res)
            cursor.close() # закрываем курсор
            conn.close()
            return ans
        
        elif sh == "":
            cursor.execute(f'SELECT sh FROM ElectroDrive WHERE rn=\'{rn}\' AND flc =\'{flc}\' and Mom={Mom} AND V=\'{V}\' AND T_pov={t} AND isp = \'{isp}\';')
            res = cursor.fetchall()
            ans = sql_format(res)
            cursor.close() # закрываем курсор
            conn.close()
            return ans

    def get_classic(self, isp = "", flc = "", N="", ob = "", V = "", Mom = "", bkw = "", elpod = "", nom = ""):
        conn = psycopg2.connect(dbname='emk', user='emk_u', password='cdjkjxm', host='127.0.0.1')
        cursor =  conn.cursor()
        print("Подбор классики")
        ans = []
        if isp=="":
            cursor.execute(f'SELECT isp FROM ClassicElectroDrive;')
            res = cursor.fetchall()
            ans = sql_format(res)
            cursor.close() # закрываем курсор
            conn.close()
            return ans
        
        elif flc=="":
            cursor.execute(f'SELECT flc FROM ClassicElectroDrive WHERE isp = \'{isp}\';')
            res = cursor.fetchall()
            ans = sql_format(res)
            cursor.close() # закрываем курсор
            conn.close()
            return ans
        
        elif (N=="") and (ob  == ""):
            cursor.execute(f'SELECT N FROM ClassicElectroDrive WHERE isp = \'{isp}\' AND flc = \'{flc}\';')
            res = cursor.fetchall()
            ans = sql_format(res)
            cursor.close() # закрываем курсор
            conn.close()
            return ans
        
        elif (N=="0") and (ob  == ""):
            cursor.execute(f'SELECT ob FROM ClassicElectroDrive WHERE isp = \'{isp}\' AND flc = \'{flc}\';')
            res = cursor.fetchall()
            ans = sql_format(res)
            cursor.close() # закрываем курсор
            conn.close()
            return ans
        
        elif (N != "") and (ob  == ""):
            cursor.execute(f'SELECT ob FROM ClassicElectroDrive WHERE isp = \'{isp}\' AND flc = \'{flc}\' AND N = \'{N}\';')
            res = cursor.fetchall()
            ans = sql_format(res)
            cursor.close() # закрываем курсор
            conn.close()
            return ans
        
        elif V =="":
            request=""
            if (N == "0") or (N == ""):
                request=""
            else:
                request=f" AND N = \'{N}\'"
                
            print(f'SELECT V FROM ClassicElectroDrive WHERE isp = \'{isp}\' AND flc = \'{flc}\'  AND ob = \'{ob}\'{request};')
            cursor.execute(f'SELECT V FROM ClassicElectroDrive WHERE isp = \'{isp}\' AND flc = \'{flc}\'  AND ob = \'{ob}\'{request};')
            res = cursor.fetchall()
            ans = sql_format(res)
            cursor.close() # закрываем курсор
            conn.close()
            return ans
        
        elif Mom =="":
            request=""
            if (N == "0") or (N == ""):
                request=""
            else:
                request=f" AND N = \'{N}\'"
            cursor.execute(f'SELECT Mom FROM ClassicElectroDrive WHERE isp = \'{isp}\' AND flc = \'{flc}\'  AND ob = \'{ob}\' AND V = \'{V}\'{request};')
            res = cursor.fetchall()
            ans = sql_format(res)
            cursor.close() # закрываем курсор
            conn.close()
            return ans
        
        elif bkw =="":
            request=""
            if (N == "0") or (N == ""):
                request=""
            else:
                request=f" AND N = \'{N}\'"
            cursor.execute(f'SELECT BKV_type FROM ClassicElectroDrive WHERE isp = \'{isp}\' AND flc = \'{flc}\'  AND ob = \'{ob}\' AND V = \'{V}\' AND Mom = {Mom}{request};')
            res = cursor.fetchall()
            ans = sql_format(res)
            cursor.close() # закрываем курсор
            conn.close()
            return ans
        
        elif elpod =="":
            request=""
            if (N == "0") or (N == ""):
                request=""
            else:
                request=f" AND N = \'{N}\'"
            cursor.execute(f'SELECT elpod FROM ClassicElectroDrive WHERE isp = \'{isp}\' AND flc = \'{flc}\'  AND ob = \'{ob}\' AND V = \'{V}\' AND Mom = \'{Mom}\' AND BKV_type = \'{bkw}\'{request};')
            res = cursor.fetchall()
            ans = sql_format(res)
            cursor.close() # закрываем курсор
            conn.close()
            return ans
        
        elif nom=="":
            request=""
            if (N == "0") or (N == ""):
                request=""
            else:
                request=f" AND N = \'{N}\'"
            cursor.execute(f'SELECT modify, N_isp FROM ClassicElectroDrive WHERE isp = \'{isp}\' AND flc = \'{flc}\'  AND ob = \'{ob}\' AND V = \'{V}\' AND Mom = \'{Mom}\' AND BKV_type = \'{bkw}\' AND elpod = \'{elpod}\'{request};')
            res = cursor.fetchall()
            a=[]
            for check in res:
                md = check[0]
                if md == None:
                    md=0
                nm = check[1]
                ans.append([md, nm])
            cursor.close() # закрываем курсор
            conn.close()
            return ans

    def get_class_engin(self, isp = "", flc = "", mod = "", nom = ""):
        conn = psycopg2.connect(dbname='emk', user='emk_u', password='cdjkjxm', host='127.0.0.1')
        cursor =  conn.cursor()
        ans = []
        isp = str(isp)
        flc=str(flc)
        mod=str(mod)
        if mod == "0":
            mod='IS NULL'
        else:
            mod='='+mod
        nom=int(nom)
        print(f'SELECT (N, N, Ipusk, Inom, U, fz, Mas) FROM ClassicElectroDrive WHERE isp=\'{isp}\' AND flc=\'{flc}\' AND mod={mod} AND N_isp=\'{nom}\';')
        cursor.execute(f'SELECT (N, N, Ipusk, Inom, U, fz, Mas) FROM ClassicElectroDrive WHERE isp=\'{isp}\' AND flc=\'{flc}\' AND modify {mod} AND N_isp=\'{nom}\';')
        res = cursor.fetchone()[0].split(',')
        print(res)
        ans = [str(res[0][1:]) + " кВт", str(res[1]) + " кВт", str(res[2]) + " А", str(res[3]) + " А", str(res[4]) + " В", str(res[5]), str(res[6][:-1])]
        cursor.close() # закрываем курсор
        conn.close()
        return ans

    def get_M1(self, sh, mom, v, ob):
        sh = int(sh)
        mom = int(mom)
        v = float(v)
        ob = float(ob)
        print(f"Подбор блока М1 при схеме - {sh}, моменте - {mom}, частоте - {v}, и оборотах - {ob}.")
        #Sim = ['M', 'L', 'K', 'J', 'I', 'H', 'G', 'F', 'E', 'D'] #слева направо
        Sim = ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M'] #справа налево
        k = 2
        m = 70

        WB = load_workbook("Params.xlsx")
        sheet = WB["M1"]

        for i in range(3, 70):
            SH = sheet[f'A{i}'].value
            Mom = sheet[f'B{i}'].value
            V = float(sheet[f'C{i}'].value) 
            if (sh == SH) and (mom == Mom) and (v == V):
                k = i
                m = i+68

        for s in Sim:
            pos_min = f"{s}{k}"
            pos_max = f"{s}{m}"

            min_ob = sheet[pos_min].value
            max_ob = sheet[pos_max].value
            
            if min_ob == "-":
                 pos_min = f"{s}2"
                 min_ob = sheet[pos_min].value
            if max_ob == "-":
                 pos_max = f"{s}70"
                 max_ob = sheet[pos_max].value
            
            min_ob = toFloat(min_ob)
            max_ob = toFloat(max_ob)

            if min_ob <= ob <= max_ob:
                print(sheet[f"{s}1"].value)
                if sheet[f"{s}1"].value == 2.5: 
                    return ".2,5"
                return "." + str(sheet[f"{s}1"].value)
        
        return ""
    
    def getMark(self, arr = ["", "", "", "", "", "", "", ""]):

        conn = psycopg2.connect(dbname='emk', user='emk_u', password='cdjkjxm', host='127.0.0.1')
        cursor =  conn.cursor()

        ans = []
        mark = arr[0]
        if mark == 'ЭП4РН':
            print("Маркировка ЭП4 с РН")
            rn = arr[1]
            isp = arr[2]
            flc = arr[3]
            mom = arr[4]
            V = arr[5]
            T_pov = arr[6]
            sh = arr[7]
            if rn == '':
                cursor.execute(f'SELECT rn FROM ElectroDrive WHERE RN IS NOT NULL;')
                res = cursor.fetchall()
                ans = sql_format(res)
            elif isp == '':
                cursor.execute(f'SELECT Isp FROM ElectroDrive WHERE RN = \'{rn}\';')
                res = cursor.fetchall()
                ans = sql_format(res)
            elif flc == '':
                cursor.execute(f'SELECT flc FROM ElectroDrive WHERE RN = \'{rn}\' AND isp = \'{isp}\';')
                res = cursor.fetchall()
                ans = sql_format(res)
            elif mom == '':
                cursor.execute(f'SELECT mom FROM ElectroDrive WHERE RN = \'{rn}\' AND isp = \'{isp}\' AND flc=\'{flc}\';')
                res = cursor.fetchall()
                ans = sql_format(res)
            elif V == '':
                cursor.execute(f'SELECT V FROM ElectroDrive WHERE RN = \'{rn}\' AND isp = \'{isp}\' AND flc=\'{flc}\' AND mom=\'{mom}\';')
                res = cursor.fetchall()
                ans = sql_format(res)
            elif T_pov == '':
                cursor.execute(f'SELECT T_pov FROM ElectroDrive WHERE RN = \'{rn}\' AND isp = \'{isp}\' AND flc=\'{flc}\' AND mom=\'{mom}\' AND V=\'{V}\';')
                res = cursor.fetchall()
                ans = sql_format(res)
            elif sh == '':
                cursor.execute(f'SELECT sh FROM ElectroDrive WHERE RN = \'{rn}\' AND isp = \'{isp}\' AND flc=\'{flc}\' AND mom=\'{mom}\' AND V=\'{V}\' AND T_pov=\'{T_pov}\';')
                res = cursor.fetchall()
                ans = sql_format(res)
            
            

        elif mark == 'ЭП4':
            print("Маркировка ЭП4")
            isp = arr[1]
            flc = arr[2]
            mom = arr[3]
            V = arr[4]
            flc_type = arr[5]
            sh = arr[6]
            print('Быстры подбор {mark}')
            if isp == '':
                cursor.execute(f'SELECT Isp FROM ElectroDrive WHERE Mark=\'{mark}\';')
                res = cursor.fetchall()
                ans = sql_format(res)
            elif flc == '':
                cursor.execute(f'SELECT flc FROM ElectroDrive WHERE Mark=\'{mark}\' AND isp = \'{isp}\';')
                res = cursor.fetchall()
                ans = sql_format(res)
            elif mom == '':
                cursor.execute(f'SELECT Mom FROM ElectroDrive WHERE Mark=\'{mark}\' AND isp=\'{isp}\' AND flc = \'{flc}\';')
                res = cursor.fetchall()
                ans = sql_format(res)
            elif V == '':
                cursor.execute(f'SELECT V FROM ElectroDrive WHERE Mark=\'{mark}\' AND isp=\'{isp}\' AND flc = \'{flc}\' and Mom = \'{mom}\';')
                res = cursor.fetchall()
                ans = sql_format(res)
            elif flc_type == '':
                cursor.execute(f'SELECT flc_type FROM ElectroDrive WHERE Mark=\'{mark}\' AND isp=\'{isp}\' AND flc = \'{flc}\' and Mom = \'{mom}\' and V = \'{V}\';')
                res = cursor.fetchall()
                ans = sql_format(res)
            elif sh == '':
                cursor.execute(f'SELECT sh FROM ElectroDrive WHERE Mark=\'{mark}\' AND isp=\'{isp}\' AND flc = \'{flc}\' and Mom = \'{mom}\' and V = \'{V}\' AND flc_type = \'{flc_type}\';')
                res = cursor.fetchall()
                ans = sql_format(res)

        elif mark == 'ЭПН':
            print("Маркировка ЭПН")
            isp = str(arr[1])
            flc = str(arr[2])
            mom = arr[3]
            V = arr[4]
            if arr[5] == "":
                arr[5]=('', '') 
            U=arr[5][0]
            fz = arr[5][1]
            sh = arr[6]
            if fz == '':
                fz = '=1'
            else:
                fz = f" = \'{fz}\'"
            sh = str(arr[6])
            if isp == '':
                cursor.execute(f'SELECT Isp FROM ElectroDrive WHERE Mark=\'{mark}\';')
                res = cursor.fetchall()
                ans = sql_format(res)
            elif flc == '':
                cursor.execute(f'SELECT flc FROM ElectroDrive WHERE Mark=\'{mark}\' AND isp=\'{isp}\';')
                res = cursor.fetchall()
                ans = sql_format(res)
            elif mom == '':
                cursor.execute(f'SELECT Mom FROM ElectroDrive WHERE Mark=\'{mark}\' AND isp=\'{isp}\' AND flc = \'{flc}\';')
                res = cursor.fetchall()
                ans = sql_format(res)
            elif V == '':
                cursor.execute(f'SELECT V FROM ElectroDrive WHERE Mark=\'{mark}\' AND isp=\'{isp}\' AND flc = \'{flc}\' and Mom = \'{mom}\';')
                res = cursor.fetchall()
                ans = sql_format(res)
            elif U == '':
                
                print(f'SELECT U FROM ElectroDrive WHERE Mark=\'{mark}\' AND Isp=\'{isp}\'  AND Mom={mom} AND V=\'{V}\' AND flc=\'{flc}\';')
                cursor.execute(f'SELECT U, fz FROM ElectroDrive WHERE Mark=\'{mark}\' AND Isp=\'{isp}\'  AND Mom={mom} AND V=\'{V}\' AND flc=\'{flc}\';')
                
                res = cursor.fetchall()
                    
                print(res)
                for a in res:
                    U=int(a[0])
                    if U == 220:
                        a = f"{U} В {a[1]} фаз(а/ы)"
                    elif  U == 380:
                        a = "380 В 3 фазы"
                    elif U == 24:
                        a = "24 В"
                    if a not in ans:
                        ans.append(a)
            elif sh == '':
                print(f'SELECT sh FROM ElectroDrive WHERE Mark=\'{mark}\' AND isp=\'{isp}\' AND flc = \'{flc}\' AND Mom = \'{mom}\' AND V = \'{V}\' AND U = \'{U}\' AND fz {fz};')
                cursor.execute(f'SELECT sh FROM ElectroDrive WHERE Mark=\'{mark}\' AND isp=\'{isp}\' AND flc = \'{flc}\' AND Mom = \'{mom}\' AND V = \'{V}\' AND U = \'{U}\' AND fz {fz};')
                res = cursor.fetchall()
                ans = sql_format(res)

        else:
            print("Маркировка Классики")
            isp = arr[1]
            flc = arr[2]
            modify = arr[3]
            N_isp = arr[4]
            key = arr[5]
            
            if isp=="":
                cursor.execute(f'SELECT isp FROM ClassicElectroDrive;')
                res = cursor.fetchall()
                ans = sql_format(res)
        
            elif flc=="":
                cursor.execute(f'SELECT flc FROM ClassicElectroDrive WHERE isp = \'{isp}\';')
                res = cursor.fetchall()
                ans = sql_format(res)
            
            elif modify=="":
                print(f'SELECT modify FROM ClassicElectroDrive WHERE isp = \'{isp}\' AND flc = \'{flc}\';')
                cursor.execute(f'SELECT modify FROM ClassicElectroDrive WHERE isp = \'{isp}\' AND flc = \'{flc}\';')
                res = cursor.fetchall()
                for mod in res:
                    if mod[0] == None:
                        print(0)
                        if 0 not in ans:
                            ans.append(0)
                    else:
                        print(mod[0])
                        if mod[0] not in ans:
                            ans.append(mod[0])
            
            
            elif N_isp == "":
                if modify == "0":
                    request = ""
                else:
                    request = f"  AND modify = \'{modify}\'"
                cursor.execute(f'SELECT N_isp FROM ClassicElectroDrive WHERE isp = \'{isp}\' AND flc = \'{flc}\'{request};')
                res = cursor.fetchall()
                ans = sql_format(res)
            
            elif key =="":
                if modify == "0":
                    request = ""
                else:
                    request = f"  AND modify = \'{modify}\'"
                cursor.execute(f'SELECT ob, V, Mom, BKV_type, elpod FROM ClassicElectroDrive WHERE isp = \'{isp}\' AND flc = \'{flc}\'{request} AND N_isp = \'{N_isp}\';')
                res = cursor.fetchall()
                ans = {
                    "Обороты" : res[0][0],
                    "Частота" : res[0][1],
                    "Крутящий момент" : res[0][2],
                    "Тип БКВ" : res[0][3],
                    "Электричеcкое подключение" : res[0][4]
                }
                

        cursor.close() # закрываем курсор
        conn.close()
        return ans



class Auth():

    def get_token(self, user: dict):
        conn = psycopg2.connect(dbname='emk', user='emk_u', password='cdjkjxm', host='127.0.0.1')
        cursor =  conn.cursor()
        token = "the data is not correct"
        #если данные валидны => даем токен
        if (len(user["fio"]) == 3):
            #если user уже есть
            cursor.execute(f'SELECT Id FROM Client WHERE name=\'{user["fio"][0]}\' AND surname=\'{user["fio"][1]}\' AND partronymic=\'{user["fio"][2]}\';')
            res = cursor.fetchone()

            
            
            #если user новый => регистрируем
            if res == None:
                cursor.execute(f"select max(id) from Client;")
                max_id = cursor.fetchone()[0]
                dep = ""
                for dr in user['department']:
                    dep= f"{dep} {dr} "
                response = f"INSERT INTO Client (ID, UUID, name, surname, partronymic, department) VALUES (\'{max_id+1}\', \'{ user['uuid'] }\', \'{user['fio'][0]}\', \'{user['fio'][1]}\', \'{user['fio'][2]}\', \'{dep}\' );"
                cursor.execute(response)

                #создаем токен
                token = encode(payload={"uuid": user["uuid"]}, key='emk', algorithm="HS512")

                print("Пользователь с ID =", max_id+1, ", создан")
            
            else:
                #создаем токен
                print("Пользователь с ID =", res[0], ", входит в систему")
                token = encode(payload={"uuid": user["uuid"]}, key='emk', algorithm="HS512")
        else:
            print("Не валидный user")
            
        cursor.close()
        conn.commit()
        conn.close()
        return token



    def get_user(self, token: str):
        print(token)
        
        #если он валидный, дать данные
        try:
            uuid = decode(token, key='emk', algorithms=["HS512"])
            valid = True
        except:
            valid = False

        #проверяем валидность токена
        user_inf = {
                "fio" : ["", "", ""], 
                "department" : "",
                "user_id" : ""
        }

        conn = psycopg2.connect(dbname='emk', user='emk_u', password='cdjkjxm', host='127.0.0.1')
        cursor =  conn.cursor()

        if valid:
            uuid = uuid["uuid"]
            user_inf["uuid"] = uuid
            cursor.execute(f"SELECT name, surname, partronymic, department, id FROM Client WHERE UUID=\'{uuid}\';")
            inf = cursor.fetchone()
            user_inf["fio"] = [inf[0], inf[1], inf[2]]
            user_inf["department"] = inf[3]
            user_inf["user_id"] = inf[4]
            
        cursor.close()
        conn.close()

        return {"valid" : valid, "user" : user_inf}
    


    def client_file(self, client_id, file_id):
        conn = psycopg2.connect(dbname='emk', user='emk_u', password='cdjkjxm', host='127.0.0.1')
        cursor =  conn.cursor()

        cursor.execute(f"select max(id) from Client_File;")
        max_id = cursor.fetchone()[0]
        cursor.execute(f"INSERT INTO Client_File (id, client_id, file_id) VALUES ({max_id+1}, \'{client_id}\', \'{file_id}\');")

        cursor.close()
        conn.commit()
        conn.close()

    def auth_file(self, ID, name, user_id):
        conn = psycopg2.connect(dbname='emk', user='emk_u', password='cdjkjxm', host='127.0.0.1')
        cursor =  conn.cursor()

        cursor.execute(f"select max(id) from File;")
        max_id = cursor.fetchone()[0]
        current_date = date.today()

        response = f"INSERT INTO File (ID, f_id, name, dt) VALUES ({max_id+1}, \'{ID}\', \'{name}\', \'{str(current_date)}\');"
        cursor.execute(response)

        cursor.close()
        conn.commit()
        conn.close()

        self.client_file(user_id, ID)


    
    def download_file(self, ID):
        conn = psycopg2.connect(dbname='emk', user='emk_u', password='cdjkjxm', host='127.0.0.1')
        cursor =  conn.cursor()

        cursor.execute(f"select f_id, name from File WHERE id=\'{ID}\';")
        file_inf = cursor.fetchone()
        
        iidd = file_inf[0]
        nm = file_inf[1]
        if nm[-4:] == ".pdf":
            fl_rormat = ".pdf"
        else:
            fl_rormat = nm[-5:]
        f_name = f"Tula{iidd}{fl_rormat}"
        cursor.close()
        conn.close()
        return [f_name, file_inf[1]]

    def get_files(self, ID):
        conn = psycopg2.connect(dbname='emk', user='emk_u', password='cdjkjxm', host='127.0.0.1')
        cursor =  conn.cursor()
        cursor.execute(f"select file_id from Client_File where client_id = \'{ID}\';")
        f_id = cursor.fetchall()
        files_id = []
        for fid in f_id:
            if fid[0] not in files_id:
                files_id.append(fid[0])
        
        fid_data = dict()
        ff_iiDD=[]
        for f_id in files_id:
            if f_id not in ff_iiDD:
                ff_iiDD.append(f_id)

        for f_id in ff_iiDD:
            cursor.execute(f"SELECT id, name FROM File WHERE f_id = \'{f_id}\' ORDER BY dt DESC;")
            dat = cursor.fetchall()
            for data in dat:
                if f_id in fid_data.keys():
                    fid_data[f_id].append({"id" : data[0], "text" : data[1]})
                else:
                    fid_data[f_id]=[{"id" : data[0], "text" : data[1]}]

        dt_fid=dict()
        for f_id in files_id:
            cursor.execute(f"SELECT dt FROM File WHERE f_id = \'{f_id}\' ORDER BY dt DESC;")
            dt = cursor.fetchall()
            for cur_dt in dt:
                current_date = cur_dt[0]
                print(current_date)
                cur_dt = f"{current_date.day}.{current_date.month}.{current_date.year}"
                if (cur_dt in dt_fid.keys()) and (f_id not in dt_fid[cur_dt]):
                    dt_fid[cur_dt].append(f_id)
                elif cur_dt not in dt_fid.keys():
                    print(cur_dt)
                    dt_fid[cur_dt] = [f_id]
        
        print(dt_fid, '\n')

        ans=dict()
        for dt in dt_fid.keys():
            fids = dt_fid[dt]
            data = []
            for fid in fids:
                for fd in fid_data[fid]:
                    data.append(fd)
            ans[dt] = data          
        
        cursor.close()
        conn.close()
        return ans

'''
{
    "uuid" : "c97f2043-7e8a-4b0f-9bf7-e6bfcf9fccb6",
    "fio" : ["Александр", "Тимофеев", "Анатольевич"], 
    "department" : ["ЭМК", "Дирекция по маркетингу ЭМК"]  
}
'''