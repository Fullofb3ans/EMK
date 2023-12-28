from openpyxl import load_workbook
from openpyxl import Workbook
import psycopg2



conn = psycopg2.connect(dbname='emk', user='emk_u', password='cdjkjxm', host='127.0.0.1')
cursor =  conn.cursor()



wb = load_workbook('DB.xlsx')
sheet = wb['ЭП']



def drive_exist(drive, dbName):
    request = ""
    for key in drive.keys():
        if  drive[key] != "":
            if request == "":
                request += f" {key} = {drive[key]}"
            else:
                if drive[key] == None:
                    request += f" AND {key} IS NULL "
                else:
                    request += f" AND {key} = \'{drive[key]}\' "
        else:
            print(drive[key])

    print(f'SELECT * FROM {dbName} WHERE {request};')
    cursor.execute(f'SELECT * FROM {dbName} WHERE {request};')
    res = cursor.fetchall()
    print(res)
    if res == []:
        return False
    else:
        return True

def add_drive(drive, dbName):
    columns = ""
    values=""
    if not drive_exist(drive, dbName):
        for key in drive.keys():
            if (drive[key] != 0) or (drive[key] != ""):
                if isinstance(drive[key], int) or isinstance(drive[key], float):
                    if columns == "":
                        columns += f"{key}"
                        values += f"{drive[key]}"
                    else:
                        columns += f", {key}"
                        values += f", {drive[key]}"
                elif isinstance(drive[key], str):
                    if columns == "":
                        columns += f"{key}"
                        values += f"{drive[key]}"
                    else:
                        columns += f", {key}"
                        values += f", \'{drive[key]}\'"
                else:
                    print(drive[key])
            else:
                print(drive[key])

        cursor.execute(f'INSERT INTO {dbName} ({columns}) VALUES ({values});')
        print(f'INSERT INTO {dbName} ({columns}) VALUES ({values});')
        

ans = []

ID = 0



for i in range(2, sheet.max_row+1):
    mark = sheet[f"A{i}"].value
    sh = sheet[f"B{i}"].value
    flc = sheet[f"C{i}"].value
    Mom = sheet[f"D{i}"].value
    V = sheet[f"E{i}"].value
    Mas = sheet[f"F{i}"].value
    N = sheet[f"H{i}"].value
    Ipusk = sheet[f"J{i}"].value
    Inom = sheet[f"K{i}"].value
    U = sheet[f"L{i}"].value
    fz = sheet[f"M{i}"].value
    
    flc_type = sheet[f"N{i}"].value

    if (Ipusk == ' ') or (Ipusk == ''):
        Ipusk=0

    drives = []

    drive = {
        "Id" : ID,
        "Mark" : mark,
        "Isp" : "",
        "sh" : str(sh),
        "flc" : str(flc),

        "Mom" : int(Mom),
        "V" : float(V),
        "Mas" : int(Mas),

        "N" : float(N),
        "Ipusk" : Ipusk,
        "Inom" : float(Inom),
        "U" : U,
        "fz" : int(fz),
        
        "flc_type" : None,
        "RN" : None,
        "T_pov" : None
    }

    for key in drive.keys():
        if (drive[key] == ' ') or (drive[key] == ''):
            drive[key] = None

    #ЭП4
    if mark == "ЭП4":
        drive["flc_type"] = int(flc_type)
        if isinstance(U, int) or isinstance(U, float):
            drive["U"] = int(U)
        else:
            try:
                drive["U"] = int(U)
            except:
                pass
        


        ISP = ["Н", "В", "Ш", "S", "С"]
        pre_drives = []

        for isp in ISP:
            drv = dict()
            drive["Isp"] = isp
            for dr in drive:
                drv[dr] = drive[dr]
            pre_drives.append(drv)
        
        for drv in pre_drives:
            drv["Id"] = ID
            if ((drv["Isp"] == "Н") or (drv["Isp"]  == "В"))    or   (((drv["Isp"] == "S") or (drv["Isp"]  == "Ш")) and ((int(sh) == 41) or (int(sh) == 410)))   or    ((drv["Isp"] == "С") and ((int(sh) == 41) or (int(sh) == 410) or (int(sh) == 40))):
                drives.append(drv)
                ID+=1
        
        #ЭП4РН
        RN = sheet[f"O{i}"].value
        Tpov = sheet[f"P{i}"].value
        
        if (RN != None):
            print(RN)
            drvs = []
            for drv in drives:
                drr=dict()
                for dr in drive:
                    drr[dr] = drv[dr]
                drvs.append(drr)
            for drv in drvs:
                drv["RN"] = RN
                drv["T_pov"] = int(Tpov)
                drv["Id"] = ID
                drives.append(drv)
                ID+=1
        
        

    #ЭПН
    elif mark == "ЭПН":
        ISP = ["Н", "В"]
        pre_drives = []

        for isp in ISP:
            drv = dict()
            drive["Isp"] = isp
            for dr in drive:
                drv[dr] = drive[dr]
            pre_drives.append(drv)

        for drv in pre_drives:
            drv["Id"] = ID
            drives.append(drv)
            ID+=1

        

    else:
        print("В экселе нет маркировки для строки: ", i)

    for drv in drives:
        add_drive(drv, 'ElectroDrive')
    
    '''for drv in drives:
        print(i, drv["Id"], drv['Mark'], drv["sh"], drv["Isp"], drv["RN"])
    print()'''

#Классика
Sheet = wb['Классика']

for i in range(2, Sheet.max_row+1):
    print(i)
    ID = i
    isp = Sheet[f"A{i}"].value
    flc = Sheet[f"B{i}"].value
    
    mod = Sheet[f"C{i}"].value
    n_isp = Sheet[f"D{i}"].value

    Mom = Sheet[f"E{i}"].value
    V = Sheet[f"F{i}"].value
    ob = Sheet[f"G{i}"].value
    Mas = Sheet[f"H{i}"].value

    N = Sheet[f"I{i}"].value
    Ipusk = Sheet[f"K{i}"].value
    Inom = Sheet[f"L{i}"].value
    U = Sheet[f"U{i}"].value
    fz = Sheet[f"N{i}"].value
    
    BKV_type = Sheet[f"O{i}"].value
    elpod = Sheet[f"P{i}"].value

    if mod == 0:
        mod = None

    drive = {
        "Id" : ID,
        "Isp" : str(isp),
        "flc" : str(flc),
        "Modify" : mod,
        "N_isp" : int(n_isp),

        "Mom" : Mom,#int(Mom),
        "V" : float(V),
        "Ob" : int(ob),
        "Mas" : int(Mas),

        "N" : float(N),
        "Ipusk" : Ipusk,
        "Inom" : float(Inom),
        "U" : 380,#int(u)
        "fz" : 3,#int(fz),
        
        "BKV_type" : BKV_type,
        "elpod" : elpod
    }

    #print(drive)
    add_drive(drive, 'ClassicElectroDrive')

#cursor.execute("DELETE FROM ElectroDrive WHERE sh='2' AND isp = 'Н';")
#cursor.execute("DELETE FROM ElectroDrive WHERE sh='31' AND Mom=300 AND V='3.0' AND isp='В';")

conn.commit()
cursor.close() # закрываем курсор
conn.close()